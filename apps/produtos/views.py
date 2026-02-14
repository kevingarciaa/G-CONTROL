from decimal import Decimal

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from .models import Produto
from .forms import ProdutoForm


def _decimal(val):
    if val is None:
        return Decimal('0')
    return Decimal(str(val))


def _custo_total_produto(cleaned_data):
    """Calcula o Custo total (soma dos blocos de cálculo) para gravar em preco_custo."""
    tecido = _decimal(cleaned_data.get('gramas_por_metro')) * _decimal(cleaned_data.get('valor_tecido'))
    est = _decimal(cleaned_data.get('quantidade_estamparia')) * _decimal(cleaned_data.get('valor_unitario_estamparia'))
    acab = _decimal(cleaned_data.get('quantidade_acabamento')) * _decimal(cleaned_data.get('valor_unitario_acabamento'))
    avi = _decimal(cleaned_data.get('quantidade_aviamentos')) * _decimal(cleaned_data.get('valor_unitario_aviamentos'))
    cost = _decimal(cleaned_data.get('quantidade_costura')) * _decimal(cleaned_data.get('valor_unitario_costura'))
    out = _decimal(cleaned_data.get('quantidade_outros')) * _decimal(cleaned_data.get('valor_unitario_outros'))
    return (tecido + est + acab + avi + cost + out).quantize(Decimal('0.01'))


def _empresa(request):
    return getattr(request.user, 'empresa', None)


@login_required
def lista(request):
    """Lista de produtos da empresa."""
    empresa = _empresa(request)
    if not empresa:
        return redirect('logout')
    produtos = Produto.objects.filter(empresa=empresa).order_by('nome')

    # Barra de pesquisa
    busca = (request.GET.get('q') or '').strip()
    if busca:
        produtos = produtos.filter(
            Q(nome__icontains=busca) |
            Q(codigo_barras__icontains=busca)
        )

    return render(request, 'produtos/lista.html', {'produtos': produtos, 'busca': busca})


@login_required
def produto_excluir_varios(request):
    """Exclui vários produtos selecionados."""
    empresa = _empresa(request)
    if not empresa:
        return redirect('logout')
    if request.method != 'POST':
        return redirect('produtos:lista')

    ids = request.POST.getlist('ids')
    excluidos = 0
    nao_excluidos = []
    for pk in ids:
        try:
            produto = Produto.objects.get(pk=pk, empresa=empresa)
            try:
                produto.delete()
                excluidos += 1
            except ProtectedError:
                nao_excluidos.append(produto.nome)
        except Produto.DoesNotExist:
            pass

    if excluidos:
        messages.success(request, f'{excluidos} produto(s) excluído(s).')
    if nao_excluidos:
        messages.warning(request, f'Não foi possível excluir: {", ".join(nao_excluidos[:3])}{"..." if len(nao_excluidos) > 3 else ""}')
    return redirect('produtos:lista')


@login_required
def produto_criar(request):
    """Cria um novo produto."""
    empresa = _empresa(request)
    if not empresa:
        return redirect('logout')
    if request.method == 'POST':
        form = ProdutoForm(request.POST, empresa=empresa)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.empresa = empresa
            obj.preco_custo = _custo_total_produto(form.cleaned_data)
            obj.save()
            messages.success(request, f'Produto "{obj.nome}" cadastrado com sucesso.')
            return redirect('produtos:lista')
    else:
        form = ProdutoForm(empresa=empresa)
    return render(request, 'produtos/produto_form.html', {'form': form, 'titulo': 'Novo produto'})


@login_required
def produto_editar(request, pk):
    """Edita um produto existente."""
    empresa = _empresa(request)
    if not empresa:
        return redirect('logout')
    produto = get_object_or_404(Produto, pk=pk, empresa=empresa)
    if request.method == 'POST':
        form = ProdutoForm(request.POST, instance=produto, empresa=empresa)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.preco_custo = _custo_total_produto(form.cleaned_data)
            obj.save()
            messages.success(request, f'Produto "{produto.nome}" atualizado.')
            return redirect('produtos:lista')
    else:
        form = ProdutoForm(instance=produto, empresa=empresa)
    return render(request, 'produtos/produto_form.html', {'form': form, 'produto': produto, 'titulo': 'Editar produto'})


@login_required
def produto_excluir(request, pk):
    """Exclui um produto."""
    empresa = _empresa(request)
    if not empresa:
        return redirect('logout')
    produto = get_object_or_404(Produto, pk=pk, empresa=empresa)
    if request.method == 'POST':
        nome = produto.nome
        produto.delete()
        messages.success(request, f'Produto "{nome}" excluído.')
        return redirect('produtos:lista')
    return render(request, 'produtos/produto_confirmar_exclusao.html', {'produto': produto})


@login_required
def estoque(request):
    """Controle de estoque: lista produtos com estoque atual e permite ajustes."""
    empresa = _empresa(request)
    if not empresa:
        return redirect('logout')
    produtos = Produto.objects.filter(empresa=empresa, ativo=True).order_by('nome')
    return render(request, 'produtos/estoque.html', {'produtos': produtos})


@login_required
def estoque_exportar_excel(request):
    """Exporta estoque de produtos para Excel, separado por menor e maior estoque."""
    empresa = _empresa(request)
    if not empresa:
        return redirect('logout')

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        from django.contrib import messages
        messages.error(request, 'Biblioteca openpyxl não instalada. Execute: pip install openpyxl')
        return redirect('produtos:estoque')

    produtos = Produto.objects.filter(empresa=empresa, ativo=True).order_by('nome')
    menor = produtos.order_by('estoque_atual')[:500]
    maior = produtos.order_by('-estoque_atual')[:500]

    wb = Workbook()
    ws_menor = wb.active
    ws_menor.title = 'Menor estoque'
    ws_maior = wb.create_sheet(title='Maior estoque')

    headers = ['Nome', 'Código barras', 'Estoque', 'Unidade', 'Preço venda']
    header_font = Font(bold=True)

    for row, ws in [(1, ws_menor), (1, ws_maior)]:
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

    for i, p in enumerate(menor, 2):
        ws_menor.cell(row=i, column=1, value=p.nome)
        ws_menor.cell(row=i, column=2, value=p.codigo_barras or '')
        ws_menor.cell(row=i, column=3, value=float(p.estoque_atual))
        ws_menor.cell(row=i, column=4, value=p.unidade)
        ws_menor.cell(row=i, column=5, value=float(p.preco_venda))

    for i, p in enumerate(maior, 2):
        ws_maior.cell(row=i, column=1, value=p.nome)
        ws_maior.cell(row=i, column=2, value=p.codigo_barras or '')
        ws_maior.cell(row=i, column=3, value=float(p.estoque_atual))
        ws_maior.cell(row=i, column=4, value=p.unidade)
        ws_maior.cell(row=i, column=5, value=float(p.preco_venda))

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="estoque_produtos.xlsx"'
    wb.save(response)
    return response


@login_required
def estoque_ajustar(request, pk):
    """Ajusta estoque de um produto (entrada ou saída)."""
    empresa = _empresa(request)
    if not empresa:
        return redirect('logout')
    produto = get_object_or_404(Produto, pk=pk, empresa=empresa)
    if request.method == 'POST':
        tipo = request.POST.get('tipo')  # 'entrada' ou 'saida'
        qtd_str = request.POST.get('quantidade', '0').replace(',', '.')
        motivo = request.POST.get('motivo', '')
        try:
            qtd = abs(float(qtd_str))
            if qtd <= 0:
                raise ValueError('Quantidade inválida')
        except (ValueError, TypeError):
            messages.error(request, 'Informe uma quantidade válida.')
            return redirect('produtos:estoque')
        if tipo == 'entrada':
            produto.estoque_atual += qtd
            msg = f'+{qtd} unidade(s) adicionada(s) ao estoque de "{produto.nome}".'
        else:
            if produto.estoque_atual < qtd:
                messages.error(request, f'Estoque insuficiente. Disponível: {produto.estoque_atual}')
                return redirect('produtos:estoque')
            produto.estoque_atual -= qtd
            msg = f'-{qtd} unidade(s) removida(s) do estoque de "{produto.nome}".'
        produto.save(update_fields=['estoque_atual'])
        messages.success(request, msg)
    return redirect('produtos:estoque')

